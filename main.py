from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from selenium import webdriver
from bs4 import BeautifulSoup
import steam.webauth as wa
from datetime import date
from requests import *
from decimal import *
import json
import time
import re


dataFolderPath = "./data/"
ScrapeDataPath = dataFolderPath + "ScrapeData/"
SteamAuthData = dataFolderPath + "SteamAuthData/"

status = None
currentGameData = list()

class ConsoleMessages:
    colours = {
        'HEADER': '\033[95m',
        'OKBLUE': '\033[94m',
        'INFOCYAN': '\033[96m',
        'INFOMAGENTA': '\033[35m',
        'OKGREEN': '\033[92m',
        'WARNING': '\033[93m',
        'FAIL': '\033[91m',
        'ENDC': '\033[0m',
        'BOLD': '\033[1m',
        'UNDERLINE': '\033[4m'}

    @staticmethod
    def input_error(): print(ConsoleMessages.colours['FAIL'] + "[FAIL]:Invalid input!" + ConsoleMessages.colours['ENDC'] + '\n')
    @staticmethod
    def no_game_data_found(): print(ConsoleMessages.colours['WARNING'] + "[WARN]:No data detected! Create or load existing data to proceed." + ConsoleMessages.colours['ENDC'] + '\n')
    @staticmethod
    def generic_success(): print(ConsoleMessages.colours['OKGREEN'] + "[OKAY]:Success!" + ConsoleMessages.colours['ENDC'] + '\n')
    @staticmethod
    def not_implemented(): print(ConsoleMessages.colours['INFOCYAN'] + "[INFO]:Not implemented!" + ConsoleMessages.colours['ENDC'] + '\n')

    @staticmethod
    def confirm_prompt(): print("Are you sure?\n"
                                + ConsoleMessages.colours['OKGREEN'] + "Y"
                                + ConsoleMessages.colours['ENDC'] + "/"
                                + ConsoleMessages.colours['FAIL'] + "N"
                                + ConsoleMessages.colours['ENDC'] + '\n')

    @staticmethod
    def ERRORMSG(message: str, haveKey: bool):
        if haveKey:
            print('\n' + ConsoleMessages.colours['FAIL'] + "[FAIL]:" + message + ConsoleMessages.colours['ENDC'] + '\n')
        else:
            print('\n' + ConsoleMessages.colours['FAIL'] + message + ConsoleMessages.colours['ENDC'] + '\n')
    @staticmethod
    def WARNMSG(message: str, haveKey: bool):
        if haveKey:
            print('\n' + ConsoleMessages.colours['WARNING'] + "[WARN]:" + message + ConsoleMessages.colours['ENDC'] + '\n')
        else:
            print('\n' + ConsoleMessages.colours['WARNING'] + message + ConsoleMessages.colours['ENDC'] + '\n')
    @staticmethod
    def OKMSG(message: str, haveKey: bool):
        if haveKey:
            print('\n' + ConsoleMessages.colours['OKGREEN'] + "[OKAY]:" + message + ConsoleMessages.colours['ENDC'] + '\n')
        else:
            print('\n' + ConsoleMessages.colours['OKGREEN'] + message + ConsoleMessages.colours['ENDC'] + '\n')
    @staticmethod
    def INFOMSG(message: str, haveKey: bool):
        if haveKey:
            print('\n' + ConsoleMessages.colours['INFOCYAN'] + "[INFO]:" + message + ConsoleMessages.colours['ENDC'] + '\n')
        else:
            print('\n' + ConsoleMessages.colours['INFOCYAN'] + message + ConsoleMessages.colours['ENDC'] + '\n')
class Menus:

    @staticmethod
    def displayOptions(optionsList):
        index = 1
        output = ""
        for option in optionsList:
            output += '[' + str(index) + '] >:' + option + '\n'
            index += 1
        print(output)

    @staticmethod
    def settingsMenu():  # TODO: Implement
        ConsoleMessages.not_implemented()

    @staticmethod
    def confirmPromptMenu():
        ConsoleMessages.confirm_prompt()

        while True:
            userInput = readInput('').lower()
            match userInput.lower():
                case 'y': return True
                case 'n': return False
                case _: ConsoleMessages.input_error()

    @staticmethod
    def scrapeExportMenu():
        # TODO: Add 'Finish' option
        # TODO: Add 'Sa'

        options = ['Export as JSON', 'Export as XLSX', 'Back']
        while True:
            ConsoleMessages.INFOMSG("Data is cached and only cleared when application is closed", True)
            print('Data path: [' + dataFolderPath + ']')
            Menus.displayOptions(options)
            match readInput(''):
                case '1':
                    GameDataScraper.convertToJSON()
                case '2':
                    GameDataScraper.convertToExcel()
                case '3':
                    return
                case _:
                    ConsoleMessages.input_error()

    @staticmethod
    def scrapeMenu():
        options = ['New scrape Event', 'Recent scrape event results', 'Back']

        while True:
            Menus.displayOptions(options)
            match readInput(''):
                case '1':
                    GameDataScraper.scrapeGameData()
                case '2':
                    if currentGameData is not None and len(currentGameData) > 0:
                        Menus.scrapeExportMenu()
                        return
                    GameDataScraper.showLastScrape()
                case '3':
                    return
                case _:
                    ConsoleMessages.input_error()

    @staticmethod
    def mainMenu():
        options = ['WebAuth', 'Scrape', 'Settings', 'Exit']

        while True:
            Menus.displayOptions(options)
            match readInput(''):
                case '1':
                    Menus.webAuthMenu()
                case '2':
                    Menus.scrapeMenu()
                case '3':
                    Menus.settingsMenu()
                case '4':
                    if Menus.confirmPromptMenu():
                        print("Exiting Application...")
                        return
                case _:
                    ConsoleMessages.input_error()

    @staticmethod
    def webAuthMenu():  # TODO: Implement
        options = ['Log in', 'Check Status', 'Back']
        ConsoleMessages.not_implemented()
class GameDataScraper:

    @staticmethod
    def __validateSteamId(steamId):
        """Checks whether the steamId leads to a valid url
        Returns game_list if website can be parsed correctly, otherwise returns None"""
        # TODO: Add an info panel that displays chosen user's details to confirm it's the correct user
        # TODO: Since we are trying to look for only a url and the div for game.ListRow
        #  static webpage way of scraping data using requests can be used to check if url is valid
        url = "https://steamcommunity.com/id/" + steamId + "/games/?tab=all&sort=playtime"
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
        driver.get(url)

        html = BeautifulSoup(driver.page_source, features='html.parser')
        game_list_div = html.select("div.gameListRow")

        if len(game_list_div) <= 0:
            ConsoleMessages.ERRORMSG("SteamId not found or profile is private!", True)
            return None

        global currentScrapedSteamId
        currentScrapedSteamId = steamId
        ConsoleMessages.OKMSG("SteamId found!", True)
        return game_list_div


    @staticmethod
    def scrapeHistory():
        ConsoleMessages.not_implemented()

    @staticmethod
    def scrapeGameData():
        # TODO: Add a way of creating a new data file[JSON or EXCEL] based on steam id provided. Each steam id will have its folder with data
        # TODO: Files should follow the following naming convention. [steamId_dd-mm-yyyy] date represents date of the scrape
        # TODO: Save previously searched users/steamids in a JSON file but only if they are valid, i.e. return a valid url
        # TODO: Create a history of recent scrape processes
        steamId = readInput("Enter steam id you wish to scrape from")

        if currentGameData is not None and len(currentGameData) > 0:
            ConsoleMessages.WARNMSG("Data already exists!\n", True)
            if not Menus.confirmPromptMenu(): return

        ConsoleMessages.INFOMSG("Scraping process initiated. Do not close Chrome process!", True)

        game_list = GameDataScraper.__validateSteamId(steamId)
        if game_list is None: return
        timeStart = time.perf_counter()

        index = 0
        all_games = []
        for game in game_list: # This code assumes the structure of the html will never change and so is prone to failing when a change occurs
            index += 1
            game_name = game.contents[2].contents[1].contents[1].contents[1].text

            try:
                game_play_time = game.contents[2].contents[1].contents[1].contents[3].contents[0].text.removesuffix(
                    ' hrs on record')
                game_play_time = re.sub(',', '', game_play_time)
            except IndexError:
                game_play_time = '0'

            all_games.append([game_name, game_play_time])

        timeEnd = time.perf_counter()

        ConsoleMessages.INFOMSG("Scraping process finished!", True)
        if index <= 0:  # TODO: Add a way of knowing if the target user's profile is set to private. Use html as a reference?
            ConsoleMessages.ERRORMSG("No games found!", True)
            return
        ConsoleMessages.OKMSG("[Games found:- "
                              + str(index) + " ]|[ Took:- "
                              + str(round(timeStart-timeEnd, 2))
                              + " seconds]", False)
        currentGameData.append(all_games)
        Menus.scrapeExportMenu()

    @staticmethod
    def showLastScrape():
        if currentGameData is None or len(currentGameData) <= 0:
            ConsoleMessages.no_game_data_found()
            return

        index = 0
        for game in currentGameData[0]:  # TODO: make currentGameData have all game data and read recent from json
            print('[' + str(index) + '] ' + game[0] + '|' + str(game[1]))
            index += 1
        print('\n')

    @staticmethod
    def convertToJSON():  #TODO: Check if file with specified name already exists and whether the user wants to overwrite it later on add feature that asks user if they want to save as new file e.g. steamId_dd_mm_yyyy(1) or overwrite
        if len(currentGameData) <= 0:
            ConsoleMessages.no_game_data_found()
            return False

        games = dict(currentGameData[0])
        games_json = json.dumps(games, indent=2)

        f = open(ScrapeDataPath + GameDataScraper.applyFileName() + ".json", "w")
        f.write(games_json)
        f = open(ScrapeDataPath + GameDataScraper.applyFileName() + ".json", "r")
        print(f.read())
        return True

    @staticmethod
    def applyFileName():
        global currentDate
        currentDate = date.today().strftime("%d-%m-%Y")
        return currentScrapedSteamId + '_' + currentDate

    @staticmethod
    def convertToExcel():
        if len(currentGameData) <= 0:
            ConsoleMessages.no_game_data_found()
            return False

        wb_name = ScrapeDataPath + GameDataScraper.applyFileName() + '.xlsx'
        wb = load_workbook(wb_name)
        ws = wb.active

        row = 2

        for game in currentGameData:
            ws['B' + str(row)] = game[0]
            ws['C' + str(row + 1)] = float(game[1])
            row += 1
        # ws.column_dimensions['C'].number_format = u'#,##0.00€'

        wb.save(wb_name)
        wb.close()
        return True

def parseLoginInfo():  # TODO: Implement input validation
    """Reads user input and parses it into username and password strings"""

    details = input("Enter your steam username and password separated by a '|' symbol\n")
    print("You entered [" + details + "]")
    return details.split('|')

def start():

    print("-:Steam Game Time Scrapper:-")
    Menus.mainMenu()

def login():
    accountDetails = parseLoginInfo()
    user = wa.WebAuth(accountDetails[0])  # Enters the username

    try:
        user.login(password=accountDetails[1])
    except (wa.LoginIncorrect, wa.CaptchaRequired, wa.TwoFactorCodeRequired) as exp:
        if isinstance(exp, wa.LoginIncorrect):
            print('Password incorrect!')
        if isinstance(exp, wa.CaptchaRequired):
            print(user.captcha_url)
            captcha = readInput("Enter Captcha code: ")
            user.login(password=accountDetails[1], captcha=captcha)
        if isinstance(exp, wa.TwoFactorCodeRequired):
            print("Steam Guard Authentication code required!")
            code = readInput("Steam Guard Code: ")
            user.login(password=accountDetails[1], twofactor_code=code)

    user.session.get('https://store.steampowered.com/account/history/')

    ConsoleMessages.OKMSG("Log in successful!", True) \
        if user.logged_on \
        else ConsoleMessages.ERRORMSG("Log in failed!", True)
    return user

def showStatus(user):
    if user is None or user.logged_on is False:
        print("Not currently signed in")
        return
    else:
        print("Logged in as [" + str(user.steam_id) + "]")
        print("Session ID [" + str(user.session_id) + "]")

def readInput(prompt):
    user_input = input(prompt + '\n>>> ')
    return user_input

if __name__ == "__main__":
    start()
